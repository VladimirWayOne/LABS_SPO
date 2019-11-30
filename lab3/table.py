import openpyxl as opxl
import numpy as np
import plot_percent
import xlrd
table = opxl.load_workbook(filename='report.xlsx')
df = table['Лист1']

Beglov = []     # процент голосов за Беглова
Tihonova = []   # процент голосов за Тихонову
Amosov = []     # процент голосов за Амосова
validSheets =[]
ChisloIzbirateley = []
VoicesBeglov=[]
for i in range(4, 79):                  # Переносим данные из таблички в массивы
    Beglov.append(df.cell(row=24, column=i).value)
    Beglov[len(Beglov)-1] = np.float(Beglov[len(Beglov)-1].replace('%',''))
    Amosov.append(df.cell(row=22, column=i).value)
    Amosov[len(Amosov)-1] = np.float(Amosov[len(Amosov)-1].replace('%', ''))
    Tihonova.append(df.cell(row=26, column=i).value)
    Tihonova[len(Tihonova)-1] = np.float(Tihonova[len(Tihonova)-1].replace('%', ''))
    validSheets.append(np.float(df.cell(row=17, column=i).value))
    ChisloIzbirateley.append(np.float(df.cell(row=9,column=i).value))
    VoicesBeglov.append(np.float(df.cell(row=23,column=i).value))

plot_percent.plot(validSheets, Beglov, 'Беглов')
plot_percent.plot(validSheets, Amosov, 'Амосов')
plot_percent.plot(validSheets, Tihonova, 'Тихонова')







